import os
import json
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import time
import re
from openpyxl import Workbook, load_workbook

# Read configuration from config.json
with open("config.json", "r") as config_file:
    config = json.load(config_file)

# Extract settings from configuration
locations = config.get("locations", [])
categories = config.get("categories", [])
page_count = config.get("page_count", 1)

# Set up the WebDriver (adjust the path to your chromedriver)
service = Service('C:/webdrivers/chromedriver.exe')
driver = webdriver.Chrome(service=service)

# Create a folder to store images
if not os.path.exists('eventbrite_images'):
    os.makedirs('eventbrite_images')

# Helper function to download the image and save it locally
def download_image(image_url, filename):
    try:
        response = requests.get(image_url, stream=True)
        if response.status_code == 200:
            image_path = os.path.join('eventbrite_images', filename)
            with open(image_path, 'wb') as f:
                f.write(response.content)
        else:
            print(f"Failed to download image: {image_url}")
    except Exception as e:
        print(f"Error downloading image: {e}")

# Helper function to sanitize filenames for Windows
def sanitize_filename(filename):
    return re.sub(r'[\/:*?"<>|]', '_', filename)

# Function to append event data to an Excel file incrementally
def append_to_excel(data, excel_file):
    if not os.path.exists(excel_file):
        # Create a new workbook and write the header and first row
        wb = Workbook()
        ws = wb.active
        ws.append(list(data.keys()))  # header row
        ws.append(list(data.values()))
        wb.save(excel_file)
    else:
        # Load the existing workbook, append the row, and save it again
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.append(list(data.values()))
        wb.save(excel_file)

# Define the Excel file path
excel_file_path = 'eventbrite_data_new.xlsx'
# (Optional) Remove the file for a fresh start:
# if os.path.exists(excel_file_path): os.remove(excel_file_path)

# Optional: List to hold all events (if you need to work with the complete data later)
events_data = []

# Set to track processed event links (to avoid duplicates)
processed_event_links = set()

# Loop through each location, category, and the specified page count
for location in locations:
    country = location.get("country")
    cities = location.get("cities", [])
    for city in cities:
        for category in categories:
            for page_number in range(1, page_count + 1):
                # Construct the URL based on country, city, category, and page number
                url = f"https://www.eventbrite.com/d/{country}--{city}/{category}--events/?page={page_number}"
                print(f"Scraping URL: {url}")
                driver.get(url)
                time.sleep(5)  # Allow time for the page to load

                # Parse the main page for event cards
                main_soup = BeautifulSoup(driver.page_source, 'html.parser')
                event_cards = main_soup.find_all('div', class_='event-card')
                print(f"Found {len(event_cards)} event cards on page {page_number} for {city} ({country}) in category '{category}'.")

                # If no event cards are found, move on to the next URL
                if not event_cards:
                    print("No events found on this page, moving to next URL.")
                    break

                # Process each event card
                for card in event_cards:
                    try:
                        min_ticket_price = card.find('div', class_='DiscoverHorizontalEventCard-module__priceWrapper___3rOUY').get_text(strip=True)

                        # Get the event link from the card
                        link_tag = card.find('a')
                        if not link_tag:
                            continue  # Skip if no link is found

                        event_link = link_tag.get('href')
                        if event_link in processed_event_links:
                            continue  # Skip duplicate event
                        processed_event_links.add(event_link)

                        # Open the event page for further details
                        driver.get(event_link)
                        time.sleep(3)  # Adjust sleep time as needed

                        # Parse the event detail page
                        event_soup = BeautifulSoup(driver.page_source, 'html.parser')

                        # Extract event name and date/time
                        event_name = event_soup.find('h1').get_text(strip=True) if event_soup.find('h1') else "N/A"
                        date_time_element = event_soup.find('span', class_='date-info__full-datetime')
                        date_time = date_time_element.get_text(strip=True) if date_time_element else "N/A"

                        # Create a sanitized filename for the event thumbnail image
                        sanitized_name = sanitize_filename(event_name)
                        sanitized_date_time = sanitize_filename(date_time.replace(' ', '_'))
                        filename = f"{sanitized_date_time}_{sanitized_name}.jpg"

                        # Extract thumbnail URL and download the image
                        thumbnail_meta = event_soup.find("meta", property="og:image")
                        if thumbnail_meta and thumbnail_meta.get("content"):
                            thumbnail_url = thumbnail_meta["content"]
                            download_image(thumbnail_url, filename)
                        else:
                            print("Thumbnail URL not found.")

                        # Extract event description
                        description = "N/A"
                        description_div = event_soup.find('div', class_='has-user-generated-content event-description')
                        if description_div:
                            description = " ".join([p.get_text(strip=True) for p in description_div.find_all('p')])

                        # Extract venue and organizer information
                        venue_element = event_soup.find('div', class_='location-info__address')
                        venue = venue_element.find('p').get_text(strip=True) if venue_element else "N/A"

                        organizer_element = event_soup.find('div', class_='descriptive-organizer-info-mobile__name')
                        organizer = organizer_element.find('a').get_text(strip=True).replace('Organized by', '').strip() if organizer_element else "N/A"

                        # Build the event data dictionary, including the min ticket price we extracted earlier
                        event_data = {
                            'event_name': event_name,
                            'description': description,
                            'date_time': date_time,
                            'venue': venue,
                            'organizer': organizer,
                            'min_ticket_price': min_ticket_price,
                            'thumbnail_filename': filename,
                            'country': country,
                            'city': city,
                            'category': category,
                            'page_number': page_number
                        }
                        events_data.append(event_data)

                        # Append the event details to the Excel file immediately
                        append_to_excel(event_data, excel_file_path)
                        print(f"Appended event: {event_name}")

                    except Exception as e:
                        print(f"Error processing event: {e}")

# Close the browser after scraping is complete
driver.quit()

print("Data extraction completed and saved incrementally to", excel_file_path)
